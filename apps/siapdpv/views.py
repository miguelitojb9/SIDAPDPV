from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from django.views.generic import CreateView, ListView, View
from apps.siapdpv.forms import QuejaForm
from apps.siapdpv.models import Queja, Documento, Tramite
from django.contrib.auth.mixins import LoginRequiredMixin
from apps.customUser.models import Departamento
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from .models import Queja
from datetime import datetime
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from datetime import date


class QuejaCreateView(LoginRequiredMixin, CreateView):
    model = Queja
    form_class = QuejaForm
    template_name = 'queja/queja_create.html'  # Plantilla para el formulario
    success_url = reverse_lazy('quejas_list')  # URL a la que se redireccionará después de crear la queja

    def form_valid(self, form):
        form.instance.recurrente = self.request.user
        form.instance.municipio = self.request.user.get_municipality_display()
        response = super().form_valid(form)

        archivos = self.request.FILES.getlist('archivos')

        for archivo in archivos:
            Documento.objects.create(queja=self.object, archivo=archivo)

        # Preparar la respuesta JSON indicando el éxito del envío del formulario

        response_data = {'success': True}

        # Enviar la respuesta JSON al cliente
        return JsonResponse(response_data)


class QuejaListView(LoginRequiredMixin, ListView):
    model = Queja
    template_name = 'queja/queja_list.html'  # Plantilla para mostrar la lista de quejas
    context_object_name = 'quejas'  # Nombre del objeto de contexto para la lista de quejas

    def get_queryset(self):
        user = self.request.user

        if user.departamento:
            # Mostrar las quejas asociadas al departamento del usuario
            return Queja.objects.filter(departamento_asignado=user.departamento, tramite=None)
        elif user.role == 'client':
            # Mostrar solo las quejas del usuario actual
            return Queja.objects.filter(recurrente=user)
        elif user.role == 'admin':
            # Mostrar todas las quejas del sistema
            return Queja.objects.filter(departamento_asignado=None)
        else:
            # En caso de otros roles o usuarios sin departamento, no mostrar ninguna queja
            return Queja.objects.none()



@login_required()
def load_queja_info(request):
    queja_id = request.GET.get('queja_id')
    queja = get_object_or_404(Queja, id=queja_id)
    # Convertir el objeto Queja a un diccionario

    queja_dict = {
        'id': queja.id,
        'fecha': queja.fecha_creacion.strftime('%Y-%m-%d'),
        'titulo': queja.asunto,
        'descripcion': queja.descripcion,

    }
    try:
        documento = Documento.objects.get(queja_id=queja.id)
        queja_dict['documento'] = documento.archivo.url
    except Documento.DoesNotExist:
        queja_dict['documento'] = None

    # Renderiza un template que contenga la información de la queja en el modal
    return JsonResponse({'data': queja_dict})


@login_required()
def asignar_departamento(request):
    if request.method == 'POST':
        try:
            queja = get_object_or_404(Queja, id=request.POST.get('queja_id'))
            departamento_id = request.POST.get('depto_id')
            departamento = get_object_or_404(Departamento, id=departamento_id)
            queja.departamento_asignado = departamento
            numero_quejas = Queja.objects.all().count() + 1
            numero_quejas_padded = str(numero_quejas).zfill(3)
            queja.noRadicacion = \
                f"{departamento.noRad}-{(numero_quejas_padded)}"

            queja.save()
            return JsonResponse({'success': True})
        except:
            return JsonResponse({'success': False})

    return JsonResponse({'success': False})


class CrearTramiteView(LoginRequiredMixin, View):
    def get(self, request, queja_id):
        queja = Queja.objects.get(id=queja_id)
        queja = Queja.objects.get(id=queja_id)
        tramite = Tramite.objects.create(
            departamento_asignado=queja.departamento_asignado,
            aprobado=True,
            tipo_tramite='I',
            usuario_responsable=request.user
        )
        queja.enTramite = True
        queja.tramite = tramite
        queja.save()
        tramite.codificacion = f"{tramite.tipo_tramite}-{queja.asunto}"
        tramite.save()
        return redirect(reverse_lazy('lista_tramites'))
        # Reemplaza con el nombre de tu URL para el listado de trámites


class CrearRespuestaTramiteView(LoginRequiredMixin, View):
    def post(self, request):
        # Obtener los datos del formulario
        tramite_id = request.POST.get('tramiteId')
        respuesta = request.POST.get('respuesta')
        conclusion_queja = request.POST.get('conclusion_queja')
        nivel_solucion = request.POST.get('nivel_solucion')

        # Obtener el archivo adjunto
        adjunto = request.FILES.get('adjunto')

        try:
            # Obtener el trámite a actualizar
            tramite = Tramite.objects.get(id=tramite_id)

            # Actualizar los datos del trámite
            tramite.respuesta = respuesta
            tramite.conclusion_queja = conclusion_queja
            tramite.nivel_solucion = nivel_solucion
            tramite.tiene_respuesta = True

            # Actualizar el archivo adjunto del trámite
            if adjunto:
                tramite.adjunto = adjunto

            # Guardar los cambios en el trámite
            queja =  Queja.objects.get(id=tramite.queja_set.first().id)
            queja.procesado = True
            queja.save()
            tramite.save()

            # Devolver una respuesta JSON
            response_data = {
                'success': True,
                'message': 'El trámite se ha actualizado correctamente.'
            }
        except Tramite.DoesNotExist:
            response_data = {
                'success': False,
                'message': 'El trámite no existe.'
            }

        return JsonResponse(response_data)


class ListarTramitesDepartamentoView(ListView):
    model = Tramite
    template_name = 'tramite/listar_tramites.html'  # Reemplaza con la ruta de tu plantilla
    context_object_name = 'tramites'

    def get_queryset(self):
        user = self.request.user
        if user.departamento:
            return Tramite.objects.filter(departamento_asignado=user.departamento, usuario_responsable=user,
                                          tiene_respuesta=False)
        elif user.role == 'admin':
            return Tramite.objects.filter(tiene_respuesta=True)
        else:
            return Tramite.objects.filter(tiene_respuesta=True, queja__recurrente=user)


@login_required()
def load_tramite_info(request):
    tramiteId = request.GET.get('tramiteId')
    tramite = get_object_or_404(Tramite, id=tramiteId)
    fecha_actual = datetime.now()
    fecha_respuesta = datetime.combine(tramite.fecha_respuesta, datetime.min.time())

    # Calcula la diferencia en días
    diferencia = (fecha_respuesta - fecha_actual).days

    # Convertir el objeto Queja a un diccionario
    tramite_dict = {
        'id': tramite.id,
        'fecha': tramite.fecha_creacion.strftime('%Y-%m-%d'),
        'dias_respuesta': diferencia,
        'fecha_respuesta': fecha_respuesta.strftime('%Y-%m-%d'),

        'responsable': tramite.usuario_responsable.first_name + '' + tramite.usuario_responsable.last_name,
        'departamento': tramite.departamento_asignado.nombre,
        'respuesta': tramite.respuesta,
        'conclusion_queja': tramite.conclusion_queja,
        'nivel_solucion': tramite.nivel_solucion,

        'queja_dict': {
            'id': tramite.queja_set.first().id,
            'titulo': tramite.queja_set.first().asunto,
            'municipio': tramite.queja_set.first().municipio,
            'cliente': tramite.queja_set.first().recurrente.first_name + ' ' +
                       tramite.queja_set.first().recurrente.last_name,
            'descripcion': tramite.queja_set.first().descripcion,
            'fecha': tramite.queja_set.first().fecha_creacion.strftime('%Y-%m-%d'),

        }
    }

    # Renderiza un template que contenga la información de la queja en el modal
    return JsonResponse({'data': tramite_dict})




@login_required
def exportar_a_excel(request):
    # Obtén tus datos para exportar, por ejemplo, desde tu modelo
    datos = Tramite.objects.filter(nivel_solucion__isnull=False)

    # Crea un nuevo archivo de Excel
    libro_excel = openpyxl.Workbook()
    hoja_excel = libro_excel.active
    # Agrega información adicional al archivo Excel
    fecha_actual = date.today().strftime('%Y-%m-%d')
    nombre_empresa = "Nombre de la empresa"
    nombre_aplicacion = "Nombre de la aplicación web"
    hoja_excel.cell(row=1, column=1, value=f"Fecha:")
    hoja_excel.cell(row=1, column=2, value=f"{fecha_actual}")
    hoja_excel.cell(row=2, column=1, value=f"Empresa:")
    hoja_excel.cell(row=2, column=2, value=f"{nombre_empresa}")
    hoja_excel.cell(row=3, column=1, value=f"Aplicación:")
    hoja_excel.cell(row=3, column=2, value=f"{nombre_aplicacion}")
    hoja_excel.cell(row=4, column=2, value=f"")


    # Agrega espacio adicional al encabezado
    encabezados = ['Fecha', 'No. Rad', 'Código', 'Procedencia', 'Nombre y Apellidos', 'Descripción', 'Departamento',
                   'Fecha Término', 'Conclusión', 'Nivel de Solución']
    hoja_excel.append(encabezados)
    for i in range(1, 6):
        for celda in hoja_excel[i]:
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            hoja_excel.row_dimensions[1].height = 30
    # Crea el estilo de borde
    borde = Border(top=Side(style='thin'), bottom=Side(style='thin'))

    # Aplica el borde a la fila 4
    for celda in hoja_excel[5]:
        celda.border = borde

    # Agrega los datos a las filas y centra el texto
    for dato in datos:
        fila = [dato.fecha_creacion.strftime('%Y-%m-%d'), dato.queja_set.first().noRadicacion, dato.codificacion,
                dato.queja_set.first().municipio, f"{dato.queja_set.first().recurrente.first_name} {dato.queja_set.first().recurrente.last_name}",
                dato.queja_set.first().descripcion, dato.departamento_asignado.nombre, dato.fecha_modificacion.strftime('%Y-%m-%d'),
                dato.conclusion_queja, dato.nivel_solucion]
        hoja_excel.append(fila)
        for celda in hoja_excel[hoja_excel.max_row]:
            celda.alignment = Alignment(horizontal='center')

    # Ajusta automáticamente el ancho de las columnas en la primera fila
    for columna in hoja_excel.columns:
        max_length = 0
        columna_letra = columna[0].column_letter
        for celda in columna:
            try:
                if len(str(celda.value)) > max_length:
                    max_length = len(celda.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Ajuste de ancho más un espacio adicional
        hoja_excel.column_dimensions[columna_letra].width = adjusted_width


    # Fija la primera fila
    hoja_excel.freeze_panes = 'A6'

    # Genera el nombre del archivo con la fecha actual
    fecha_actual = date.today().strftime('%Y-%m-%d')
    nombre_archivo = f"{fecha_actual} estadistico SAPV.xlsx"

    # Crea una respuesta HTTP con el archivo Excel adjunto
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

    # Guarda el archivo Excel en la respuesta HTTP
    libro_excel.save(response)

    return response